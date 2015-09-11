import os       # library for changing directory calls
import Tkinter, tkFileDialog   # library for open and save file dialogs
from Tkinter import *
import openpyxl as px     # read xlsx library
import re

from math import radians, cos, sin, asin, sqrt

def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 

    # 6371 km is the radius of the Earth
    miles = 6371 * c * 1000 * 100 /2.54 / 12 /5280
    return miles

# make the home directory the starting point of the open file dialog
from os.path import expanduser
home = expanduser("~")     # get default directory of user
os.chdir(home)
os.chdir("./Documents")
#os.chdir('C:\Users\DDalton\Documents\RF performance\CDMA\ODD')

# Open CDMA 1x parameters spreadsheet
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='EVDO Ncarrier spreadsheet')
try:
    #read workbook
    wbncar = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    wsncar = wbncar.get_sheet_by_name(name = 'NBCARRIERINFO')
    #ws = wb.active
except:
    print "error opening file"
    exit()

# make directory of first file the new default directory
os.chdir(os.path.dirname(file))

# Open spreadsheet with accurate lat/lon info
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='CDMA spreadsheet with lat-lon info in colmns 3 and 4 and cell BTS number in column 2')
try:
    #read workbook
    wblatlon = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    wslatlon = wblatlon.get_sheet_by_name(name = 'Sheet1')
    #ws = wb.active
except:
    print "error opening lat-lon file"
    exit()

# open output file
# get saveas filename for output
root = Tkinter.Tk()
root.withdraw()
myFormats = [
    ('csv','*.csv'),
    ]

fileName = tkFileDialog.asksaveasfilename(filetypes=myFormats ,title="Save the output csv file as...")
if len(fileName ) < 1:
    print "bad file name"
    exit()
try:
    whand = open(fileName,'w')
except:
    print "bad file name"
    exit()
    
#dictionaries to hold cell and lat/lon.  Cell only, not sector
cellLat= dict()
cellLon = dict()

# first line is header so skip it
first = 1
i=0

for row in wslatlon.iter_rows():
    if first < 2:       # 1 rows of header
        first += 1
        continue
    if row[1].value < 1:    # end of list of sites
        break
    cellLat[str(row[1].value)] = str(row[2].value)  # populate dictionary with cell lat lon
    cellLon[str(row[1].value)] = str(row[3].value)

# change to nbr list sheet

first = 1
for row in wsncar.iter_rows():
    if first < 6:       # 5 rows of headers
        first += 1
        continue
    source = str(row[2].value)
    dest = str(row[7].value)
    lat1 = float(cellLat[source])
    lon1 = float(cellLon[source])
    lat2 = float(cellLat[dest])
    lon2 = float(cellLon[dest])
    dist = haversine(lon1, lat1, lon2, lat2)
    temp = row[18].value
    if dist < 4.6:
        srchwin = 8
    elif dist < 6.1:
        srchwin = 9
    elif dist < 7.6:
        srchwin = 10
    elif dist < 9.9:
        srchwin = 11
    elif dist < 12.1:
        srchwin = 12
    elif dist < 17.1:
        srchwin = 13
    else:
        srchwin = 14

    whand.write("Source," + source + ",Dest," + dest + ",dist," + str(dist) + ",SRCHWIN," + str(srchwin) + ",old," + str(temp) + "\n")
    
whand.close()
