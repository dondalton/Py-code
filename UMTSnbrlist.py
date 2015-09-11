import os       # library for changing directory calls
import Tkinter, tkFileDialog   # library for open and save file dialogs
from Tkinter import *
import openpyxl as px     # read xlsx library

from math import radians, cos, sin, asin, sqrt

def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 

    # 6371 km is the radius of the Earth
    miles = 6371 * c * 1000 * 100 /2.54 / 12 /5280
    return miles

# make the home directory the starting point of the open file dialog
from os.path import expanduser
home = expanduser("~")     # get default directory of user
os.chdir(home)
os.chdir("./Documents")
#os.chdir('C:\Users\DDalton\Documents\RF performance\UMTS')  # just to speed up testing

master = Tk()

Label(master, text="Min Distance").grid(row=0)
Label(master, text="Max Distance").grid(row=1)

e1 = Entry(master)
e2 = Entry(master)
e1.insert(10,"10")
e2.insert(10,"20")

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)

Button(master, text='Continue', command=master.quit).grid(row=3, column=0, sticky=W, pady=4)
#Button(master, text='Show', command=show_entry_fields).grid(row=3, column=1, sticky=W, pady=4)

mainloop()
max = float(e2.get())
min = float(e1.get())
master.withdraw()


root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='UMTS switch dump')
try:
    #read workbook
    wb = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    ws = wb.get_sheet_by_name(name = 'UUtranCellFDD')
    #ws = wb.active
except:
    print "error opening file"
    exit()

# make directory of first file the new default directory
os.chdir(os.path.dirname(file))

# get saveas filename for output
root = Tkinter.Tk()
root.withdraw()
myFormats = [
    ('text','*.txt'),
    ]

fileName = tkFileDialog.asksaveasfilename(filetypes=myFormats ,title="Save the output text file as...")
if len(fileName ) < 1:
    print "bad file name"
    exit()
try:
    whand = open(fileName,'w')
except:
    print "bad file name"
    exit()
    
#dictionaries to hold cell and lat/lon.  Cell only, not sector
cellLat= dict()
cellLon = dict()
# dict to hold all cell-sector names.  must strip sector too look up lat/lon
seclist = []

# first line is header so skip it
first = 1
i=0

for row in ws.iter_rows():
    if first < 6:       # 6 rows of header
        first += 1
        continue
    if float(row[55].value) > 50.0 or float(row[55].value) < 20.0 or float(row[56].value) > -60.0 or float(row[56].value) < -125:
        whand.write(row[5].value + " has a bad lat/lon value: " + row[55].value + " / " + row[56].value + '\n')
    cellid = str(row[5].value[:-1])
    cellLat[cellid] = row[55].value  # populate dictionary with cell (not sector) lat lon
    cellLon[cellid] = row[56].value
    seclist.append(row[5].value)   # list of all cellID-sector values

# change to nbr list sheet
ws = wb.get_sheet_by_name(name = 'UUtranRelation')

# list of lists, source-sect and nbr-sect
nbrlist = []
#temp holder of two part list, source cell-sect and dest cell-sect.  Will be appended to nbrlist
nbrcell = []

first = 1
for row in ws.iter_rows():
    if first < 6:       # 6 rows of headers
        first += 1
        continue
    source = str(row[5].value)
    if row[10].value != "":
        dest = row[10].value
        tempdest = dest.split(",")
        tempcell = str(tempdest[3].split("=")[1])
        nbrcell = [source,tempcell]
        nbrlist.append(nbrcell)
        
#print nbrlist

for sector in seclist:
    for destsector in seclist:
        nbrcell = [sector,destsector]
        sect = str(sector[:-1])
        destsect = str(destsector[:-1])
        lat1 = float(cellLat[sect])
        lon1 = float(cellLon[sect])
        lat2 = float(cellLat[destsect])
        lon2 = float(cellLon[destsect])
        distance = haversine(lon1, lat1, lon2, lat2)
        if sector != destsector:
            if nbrcell not in nbrlist and distance < min:
                whand.write(sector + " is missing neighbor " + destsector + " which is " + str(distance) + " miles away" + '\n')
            if nbrcell in nbrlist and distance > max:
                whand.write(sector + " has neighbor " + destsector + " that is " + str(distance) + " miles away" + '\n')
    
whand.close()
