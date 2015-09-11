import os       # library for changing directory calls
import Tkinter, tkFileDialog   # library for open and save file dialogs
from Tkinter import *
import openpyxl as px     # read xlsx library
import re

from math import radians, cos, sin, asin, sqrt

def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 

    # 6371 km is the radius of the Earth
    miles = 6371 * c * 1000 * 100 /2.54 / 12 /5280
    return miles

# make the home directory the starting point of the open file dialog
from os.path import expanduser
home = expanduser("~")     # get default directory of user
os.chdir(home)
os.chdir("./Documents")
#os.chdir('C:\Users\DDalton\Documents\RF performance\CDMA\ODD')

master = Tk()

Label(master, text="Min Distance").grid(row=0)
Label(master, text="Max Distance").grid(row=1)

e1 = Entry(master)
e2 = Entry(master)
e1.insert(10,"10")
e2.insert(10,"20")

e1.grid(row=0, column=1)
e2.grid(row=1, column=1)

Button(master, text='Continue', command=master.quit).grid(row=3, column=0, sticky=W, pady=4)

mainloop()
max = float(e2.get())
min = float(e1.get())
master.withdraw()

# Open CDMA 1x parameters spreadsheet
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='CDMA 1x Parameters spreadsheet')
try:
    #read workbook
    wb = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    ws = wb.get_sheet_by_name(name = 'CELL')
    #ws = wb.active
except:
    print "error opening file"
    exit()

# make directory of first file the new default directory
os.chdir(os.path.dirname(file))

# Open NCell&Ncarrier spreadsheet
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='CDMA 1x NGListInfo(BTS1-xxxx) spreadsheet')
try:
    #read workbook
    wbnbr = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    wsnbr = wbnbr.get_sheet_by_name(name = 'NGLISTINFO')
    #ws = wb.active
except:
    print "error opening neighborlist file"
    exit()

# get saveas filename for output
root = Tkinter.Tk()
root.withdraw()
myFormats = [
    ('text','*.txt'),
    ]

fileName = tkFileDialog.asksaveasfilename(filetypes=myFormats ,title="Save the output text file as...")
if len(fileName ) < 1:
    print "bad file name"
    exit()
try:
    whand = open(fileName,'w')
except:
    print "bad file name"
    exit()
    
#dictionaries to hold cell and lat/lon.  Cell only, not sector
cellLat= dict()
cellLon = dict()
# dict to hold all cell-sector names.  must strip sector too look up lat/lon
seclist = []

# first line is header so skip it
first = 1
i=0

for row in ws.iter_rows():
    if first < 12:       # 11 rows of header
        first += 1
        continue
    if row[1].value < 1:    # end of cell parameters start of sys_para section
        break
    if row[27].value < 1 or row[29].value < 1:
        whand.write(str(row[1].value) + " has a bad lat/lon value: NULL " + '\n')
        cellLat[row[1].value] = 0  # populate dictionary with cell lat lon
        cellLon[row[1].value] = 0
        continue
    elif float(row[27].value) > 50.0 or float(row[27].value) < 20.0 or float(row[29].value) > -60.0 or float(row[29].value) < -125:
        whand.write(str(row[1].value) + " has a bad lat/lon value: " + str(row[27].value) + " / " + str(row[29].value) + '\n')
    cellLat[str(row[1].value)] = row[27].value  # populate dictionary with cell lat lon
    cellLon[str(row[1].value)] = row[29].value
    temp = str(row[1].value) + '-' + str(row[3].value)      # cellID-sector
    seclist.append(temp)   # list of all cellID-sector values

### change to nbr list sheet

# list of lists, source-sect and nbr-sect
nbrlist = []
#temp holder of two part list, source cell-sect and dest cell-sect.  Will be appended to nbrlist
nbrcell = []

first = 1
for row in wsnbr.iter_rows():
    if first < 2:       # 2 rows of headers
        first += 1
        continue
    source = str(row[1].value) + "-" + str(row[2].value)
    dest = str(row[5].value) + "-" + str(row[6].value)
    nbrcell = [source,dest]
    nbrlist.append(nbrcell)

for sector in seclist:
    for destsector in seclist:
        nbrcell = [sector,destsector]
        sect = sector.split("-")[0]
        destsect = destsector.split("-")[0]
        lat1 = float(cellLat[sect])
        lon1 = float(cellLon[sect])
        lat2 = float(cellLat[destsect])
        lon2 = float(cellLon[destsect])
        distance = haversine(lon1, lat1, lon2, lat2)
        if sector != destsector:
            if nbrcell not in nbrlist and distance < min:
                whand.write(sector + " is missing neighbor " + destsector + " which is " + str(distance) + " miles away" + '\n')
            if nbrcell in nbrlist and distance > max:
                whand.write(sector + " has neighbor " + destsector + " that is " + str(distance) + " miles away" + '\n')
    
whand.close()
