import os       # library for changing directory calls
import Tkinter, tkFileDialog   # library for open and save file dialogs
from Tkinter import *
import openpyxl as px     # read xlsx library
import re
from math import radians, cos, sin, asin, sqrt

def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians 
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula 
    dlon = lon2 - lon1 
    dlat = lat2 - lat1 
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a)) 

    # 6371 km is the radius of the Earth
    miles = 6371 * c * 1000 * 100 /2.54 / 12 /5280
    return miles

# make the home directory the starting point of the open file dialog
from os.path import expanduser
home = expanduser("~")     # get default directory of user
os.chdir(home)
os.chdir("./Documents")

# Open CDMA 1x parameters spreadsheet
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='CDMA spreadsheet')
try:
    #read workbook
    wbcdma = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    wscdma = wbcdma.get_sheet_by_name(name = 'List Sectors')
    #ws = wb.active
except:
    print "error opening file"
    exit()

# make directory of first file the new default directory
os.chdir(os.path.dirname(file))

# Open NCell&Ncarrier spreadsheet
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(title='UMTS spreadsheet')
try:
    #read workbook
    wbumts = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    wsumts = wbumts.get_sheet_by_name(name = 'List Sectors')
    #ws = wb.active
except:
    print "error opening neighborlist file"
    exit()

# get saveas filename for output
root = Tkinter.Tk()
root.withdraw()
myFormats = [
    ('text','*.txt'),
    ]

fileName = tkFileDialog.asksaveasfilename(filetypes=myFormats ,title="Save the output text file as...")
if len(fileName ) < 1:
    print "bad file name"
    exit()
try:
    whand = open(fileName,'w')
except:
    print "bad file name"
    exit()

# dictionary for cdma name[lat]
cdmalat = {}
#dictionary for cdma name[lon]
cdmalon = {}
#list of cdma sector names
cdmaname = []

# first line is header so skip it
first = 1
i=0

for row in wscdma.iter_rows():
    if first < 2:       # 1 rows of header
        first += 1
        continue
    if row[1].value < 1:    # end of data
        break
    cdmalat[str(row[0].value)] = float(row[1].value)
    cdmalon[str(row[0].value)] = float(row[2].value)
    cdmaname.append(str(row[0].value))   # list of site names

### change to umts sheet
# first line is header so skip it
first = 1
i=0

for row in wsumts.iter_rows():
    if first < 2:       # 1 rows of header
        first += 1
        continue
    if row[1].value < 1:    # end of data
        break
    umtslat = float(row[1].value)
    umtslon = float(row[2].value)
    umtsname = row[0].value
    mindist = 9999    # closest CDMA site to the UMTS site distance
    for cdmasite in cdmaname:
        lat1 = cdmalat[cdmasite]
        lon1 = cdmalon[cdmasite]
        dist = haversine(lon1,lat1,umtslon,umtslat)
        if dist < mindist:
            mindist = dist
            #print str(dist) + "," + cdmasite + "," + str(lat1) + "," + umtsname + "," + str(umtslat)

    if mindist > 0.1:
        whand.write(umtsname + "," + str(umtslat) + "," + str(umtslon)+ '\n')
        print umtsname + "," + str(umtslat) + "," + str(umtslon)+ '\n'
whand.close()
