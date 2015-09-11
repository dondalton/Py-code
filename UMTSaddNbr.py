import os       # library for changing directory calls
import Tkinter, tkFileDialog   # library for open and save file dialogs
from Tkinter import *
import openpyxl as px     # read xlsx library
from openpyxl import Workbook   # create new workbook
from openpyxl.writer.excel import ExcelWriter  # writes to excel
from openpyxl.cell import get_column_letter   

# make the home directory the starting point of the open file dialog
from os.path import expanduser
home = expanduser("~")     # get default directory of user
os.chdir(home)
os.chdir("./Documents")
os.chdir('C:\Users\DDalton\Documents\RF performance\UMTS')  # just to speed up testing

# file types to be used in dialogs
myFormats = [
    ('Excel Workbook','*.xlsx'),
    ]
#Open file dialog to get switch dump
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(filetypes=myFormats ,title='UMTS switch dump in xlsx format')
try:
    #read workbook
    wbdump = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    wsdump = wbdump.get_sheet_by_name(name = 'UUtranRelation')
    #ws = wb.active
except:
    print "error opening file:" + file
    exit()

# make directory of first file the new default directory
os.chdir(os.path.dirname(file))

#Open file dialog to get new neighbor list xlsx file
root = Tkinter.Tk()
root.withdraw()
file = tkFileDialog.askopenfilename(filetypes=myFormats ,title='UMTS neighbors to be added (*.xlsx)')
try:
    #read workbook
    wbnbr = px.load_workbook(file, use_iterators = True)
    # set ws to active workbook, active sheet
    #wsnbr = wbnbr.get_sheet_by_name(name = 'UUtranCellFDD')
    wsnbr = wbnbr.active
except:
    print "error opening file:" + file
    exit()

# make directory of first file the new default directory
os.chdir(os.path.dirname(file))

# get saveas filename for output file *.xlsx
root = Tkinter.Tk()
root.withdraw()

destFileName = tkFileDialog.asksaveasfilename(filetypes=myFormats ,title="Save the output *.xlsx file as...")
if len(destFileName ) < 1:
    print "bad file name"
    exit()

#check for .xlsx extention
if '.xlsx' not in destFileName:
    destFileName = destFileName + ".xlsx"

# make new workbook
wbnew = Workbook()
ew = ExcelWriter(workbook = wbnew)
wsnew = wbnew.worksheets[0]
wsnew.title = "UUtranRelation"
    
# copy headers from wsdump to wbnbr
rownum = 1
colnum = 1
for row in wsdump.iter_rows():
    for cell in row:
        colletter = get_column_letter(colnum)
        #print colletter, rownum, cell.value
        #print '%s%s'%(colletter,rownum)
        wsnew.cell('%s%s'%(colletter,rownum)).value = cell.value
        colnum = colnum + 1
    colnum = 1
    rownum = rownum + 1
    if rownum > 5:
        break

#save file to destFileName 
ew.save(filename = destFileName)
