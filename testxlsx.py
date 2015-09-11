import openpyxl as px
import os

os.chdir("C:\Users\DDalton\Documents\python")

#read workbook
wb = px.load_workbook('test.xlsx')
# read sheetnames
p = wb.get_sheet_by_name(name = 'Sheet1')
print wb.get_sheet_names()
# set ws to active workbook, active sheet
ws = wb.active

# print specific cell values
print ws['A1'].value
print ws.cell('A4').value

# open file with iterators enabled
wb1 = px.load_workbook('test.xlsx', use_iterators = True)
ws = wb.get_sheet_by_name(name = 'Sheet1')
# iterate through rows, then each cell in row and print value
for row in ws.iter_rows():
    for cell in row:
        print cell.value

print "done"

from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
 
from openpyxl.cell import get_column_letter
 
# make new workbook
wb = Workbook()
 
ew = ExcelWriter(workbook = wb)
 
dest_filename = r'empty_book.xlsx'
 
ws = wb.worksheets[0]
 
ws.title = "range names"

# run through 1 - 40 columns and 1 - 600 rows and fill them with row/column name 
for col_idx in xrange(1, 40):
    col = get_column_letter(col_idx)
    for row in xrange(1, 600):
        ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)
 
#create a new sheet and call it 'Pi'
ws = wb.create_sheet()
 
ws.title = 'Pi'
 
#assign a value to cell F5 on new sheet
ws.cell('F5').value = 3.14

#save file to 'empty_book.xlsx' 
ew.save(filename = dest_filename)
